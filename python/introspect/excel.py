"""Introspección Excel local: fila de headers (header_row). Todo VARCHAR. No extrae filas."""

from __future__ import annotations

from pathlib import Path

from .h2_ddl import Column, map_h2_type, sanitize_ident


def introspect(source: dict, variables: dict[str, str], root: Path | None = None) -> list[Column]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Falta openpyxl. Instala: pip install -r python/requirements.txt"
        ) from exc

    if root is None:
        raise ValueError("excel: falta root del proyecto")

    raw_path = (source.get("path") or "").strip()
    worksheet = source.get("worksheet")
    if not raw_path:
        raise ValueError(f"{source.get('stg_table')}: falta path")
    if not worksheet:
        raise ValueError(f"{source.get('stg_table')}: falta worksheet")

    path = Path(raw_path)
    if not path.is_absolute():
        path = (root / path).resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Excel no encontrado: {path}")

    header_raw = source.get("header_row", 1)
    try:
        header_row = int(header_raw)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"{source.get('stg_table')}: header_row debe ser entero, recibido {header_raw!r}"
        ) from exc
    if header_row < 1:
        raise ValueError(f"{source.get('stg_table')}: header_row debe ser >= 1")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if str(worksheet) not in wb.sheetnames:
            raise ValueError(
                f"Excel '{path.name}': no existe pestaña {worksheet!r}. "
                f"Hojas: {wb.sheetnames}"
            )
        ws = wb[str(worksheet)]
        headers = None
        for i, row in enumerate(ws.iter_rows(max_row=header_row, values_only=True), 1):
            if i == header_row:
                headers = list(row)
                break
    finally:
        wb.close()

    if not headers:
        raise ValueError(f"Excel '{worksheet}': fila {header_row} vacía")

    used: set[str] = set()
    cols: list[Column] = []
    for header in headers:
        if header is None or str(header).strip() == "":
            continue
        cols.append(
            Column(
                name=sanitize_ident(str(header), used),
                h2_type=map_h2_type("VARCHAR", sheets=True),
            )
        )
    if not cols:
        raise ValueError(f"Excel '{worksheet}': fila {header_row} sin headers usables")
    return cols
